import openpyxl

inv_file = openpyxl.load_workbook("inventory.xlsx")

product_list = inv_file["Sheet1"]

product_per_suplier = {}
total_value_per_suplier = {}
product_under_10_inv = {}

for product_row in range(2, product_list.max_row + 1):
    suplier_names = product_list.cell(product_row, 4).value
    inventory = product_list.cell(product_row, 2).value
    price = product_list.cell(product_row, 3).value
    producut_muber = product_list.cell(product_row, 1).value
    total_inv_price = product_list.cell(product_row, 5)

    # print(f"{product_row}. {suplier_names} has {inventory} invertories and total price is: {inventory*price}")
    #Calculation number of product per suplier 
    if suplier_names in product_per_suplier:
        current_num_product = product_per_suplier.get(suplier_names)
        product_per_suplier[suplier_names] = current_num_product + 1
    else:
        product_per_suplier[suplier_names] = 1

    #Calculation total value of inventory per suplier
    if suplier_names in total_value_per_suplier:
        current_total_value = total_value_per_suplier.get(suplier_names)
        total_value_per_suplier[suplier_names] = current_total_value + inventory * price
    else:
        total_value_per_suplier[suplier_names] = inventory * price

    #Product under inventory 10 
    if inventory < 50:
        product_under_10_inv[int(producut_muber)] = int(inventory)

    #Add total value of inventory in colum 5
    total_inv_price.value = inventory * price

print(product_per_suplier)
print(total_value_per_suplier)
print(product_under_10_inv)
inv_file.save("inventory2.xlsx")
